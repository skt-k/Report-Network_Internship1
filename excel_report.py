import os
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference
from openpyxl.styles import Alignment, PatternFill
from utils import (
    ALL_COLS,
    ALL_COLS_TH,
    DATA_FONT,
    HEADER_FILL,
    HEADER_FONT,
    MANUAL_BLANK_COLS,
    RATING_FILLS,
    FIELD_MAP,
    safe_float,
    safe_val,
    shade_cell,
    style_word_header,
    thin_border,
    make_bar_chart,
    make_line_chart,
)


def _format_column_widths(ws):
    for column_cells in ws.columns:
        max_width = max((len(str(cell.value or "")) for cell in column_cells), default=8) + 3
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_width, 22)


def _create_table_sheet(ws, headers, rows, extra_fill=True):
    for ci, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    for ri, row_data in enumerate(rows, start=2):
        for ci, value in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = DATA_FONT
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
            if extra_fill and ci == len(row_data) and isinstance(value, str):
                cell.fill = RATING_FILLS.get(value, PatternFill("solid", fgColor="FFFFFF"))


def create_excel_report(df, df_hops, target_building, output_dir):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"

    _create_table_sheet(
        ws_data,
        ALL_COLS_TH,
        [
            [
                float(row[col]) if col not in {"note", "building", "ssid", "bssid", "ap_vendor", "band", "radio_type", "gateway_ip", "server_ip", "trace_target", "rating"} and row.get(col) is not None else row.get(col)
                for col in ALL_COLS
            ]
            for _, row in df.iterrows()
        ],
        extra_fill=False,
    )

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for ci, col in enumerate(ALL_COLS, start=1):
            cell = ws_data.cell(row=row_idx, column=ci)
            if col == "rating":
                cell.fill = RATING_FILLS.get(str(cell.value), PatternFill("solid", fgColor="FFFFFF"))

    ws_data.row_dimensions[1].height = 35
    ws_data.freeze_panes = "A2"
    _format_column_widths(ws_data)

    # RSSI chart sheet
    ws_rssi = wb.create_sheet("กราฟ RSSI")
    ws_rssi.append(["ห้อง / จุด", "RSSI (dBm)"])
    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws_rssi.cell(row=i, column=1, value=f"{getattr(row, 'room_point')} - {getattr(row, 'note', '')}")
        ws_rssi.cell(row=i, column=2, value=safe_float(getattr(row, 'rssi_dbm')))

    ws_rssi.column_dimensions["A"].width = 30
    ws_rssi.column_dimensions["B"].width = 14

    rssi_values = df["rssi_dbm"].dropna().astype(float)
    rssi_min = int(rssi_values.min()) - 5 if len(rssi_values) else -80
    rssi_max = 0

    rssi_chart = make_bar_chart(
        title="RSSI (dBm) ตามจุดสำรวจ",
        y_title="dBm",
        x_title="ห้อง / จุดสำรวจ",
        y_min=rssi_min,
        y_max=rssi_max,
        major_unit=5,
    )
    rssi_chart.add_data(Reference(ws_rssi, min_col=2, min_row=1, max_row=len(df) + 1), titles_from_data=True)
    rssi_chart.set_categories(Reference(ws_rssi, min_col=1, min_row=2, max_row=len(df) + 1))
    ws_rssi.add_chart(rssi_chart, "E2")

    # Speed chart sheet
    ws_speed = wb.create_sheet("กราฟ Speed")
    ws_speed.append(["ห้อง", "TCP Upload (Mbps)", "TCP Download (Mbps)", "UDP Actual (Mbps)"])
    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws_speed.cell(row=i, column=1, value=f"{getattr(row, 'room_point')} - {getattr(row, 'note', '')}")
        ws_speed.cell(row=i, column=2, value=safe_float(getattr(row, 'tcp_upload_mbps')))
        ws_speed.cell(row=i, column=3, value=safe_float(getattr(row, 'tcp_download_mbps')))
        ws_speed.cell(row=i, column=4, value=safe_float(getattr(row, 'udp_actual_mbps')))

    ws_speed.column_dimensions["A"].width = 30
    speed_chart = make_bar_chart(
        title="ความเร็ว TCP Upload / Download / UDP (Mbps)",
        y_title="Mbps",
        x_title="ห้อง / จุดสำรวจ",
        y_min=0,
        width=28,
        height=15,
    )
    speed_chart.add_data(Reference(ws_speed, min_col=2, min_row=1, max_col=4, max_row=len(df) + 1), titles_from_data=True)
    speed_chart.set_categories(Reference(ws_speed, min_col=1, min_row=2, max_row=len(df) + 1))
    ws_speed.add_chart(speed_chart, "F2")

    # Latency chart sheet
    ws_lat = wb.create_sheet("กราฟ Latency")
    ws_lat.append([
        "ห้อง",
        "Ping Gateway (ms)",
        "Ping Server (ms)",
        "Gateway Loss %",
        "Server Loss %",
    ])
    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws_lat.cell(row=i, column=1, value=f"{getattr(row, 'room_point')} - {getattr(row, 'note', '')}")
        ws_lat.cell(row=i, column=2, value=safe_float(getattr(row, 'ping_gateway_ms')))
        ws_lat.cell(row=i, column=3, value=safe_float(getattr(row, 'ping_server_ms')))
        ws_lat.cell(row=i, column=4, value=safe_float(getattr(row, 'ping_gateway_loss_pct')))
        ws_lat.cell(row=i, column=5, value=safe_float(getattr(row, 'ping_server_loss_pct')))

    ws_lat.column_dimensions["A"].width = 30

    lat_values = []
    lat_values.extend(df["ping_gateway_ms"].dropna().astype(float).tolist())
    lat_values.extend(df["ping_server_ms"].dropna().astype(float).tolist())
    lat_max = int(np.percentile(lat_values, 95)) + 20 if lat_values else 100

    latency_chart = make_bar_chart(
        title="Latency — Ping Gateway vs Server (ms)",
        y_title="ms",
        x_title="ห้อง / จุดสำรวจ",
        y_min=0,
        y_max=lat_max,
        width=28,
        height=14,
    )
    latency_chart.add_data(Reference(ws_lat, min_col=2, min_row=1, max_col=3, max_row=len(df) + 1), titles_from_data=True)
    latency_chart.set_categories(Reference(ws_lat, min_col=1, min_row=2, max_row=len(df) + 1))
    ws_lat.add_chart(latency_chart, "G2")

    loss_chart = make_line_chart(
        title="Packet Loss % — Gateway & Server",
        y_title="%",
        x_title="ห้อง / จุดสำรวจ",
        y_min=0,
        width=28,
        height=14,
    )
    loss_chart.add_data(Reference(ws_lat, min_col=4, min_row=1, max_col=5, max_row=len(df) + 1), titles_from_data=True)
    loss_chart.set_categories(Reference(ws_lat, min_col=1, min_row=2, max_row=len(df) + 1))
    ws_lat.add_chart(loss_chart, "G22")

    if not df_hops.empty:
        ws_mtr = wb.create_sheet("WinMTR Data")
        mtr_cols = [
            "survey_id",
            "room_point",
            "hop_no",
            "ip",
            "loss_pct",
            "sent",
            "recv",
            "best_ms",
            "avg_ms",
            "worst_ms",
            "last_ms",
        ]
        ws_mtr.append([
            "Survey ID",
            "ห้อง",
            "Hop #",
            "IP",
            "Loss %",
            "Sent",
            "Recv",
            "Best (ms)",
            "Avg (ms)",
            "Worst (ms)",
            "Last (ms)",
        ])
        for _, row in df_hops.iterrows():
            ws_mtr.append([row.get(col) for col in mtr_cols])
        _format_column_widths(ws_mtr)
        ws_mtr.freeze_panes = "A2"

    xlsx_name = os.path.join(output_dir, f"wifi_report_{target_building}.xlsx")
    wb.save(xlsx_name)
    return xlsx_name
